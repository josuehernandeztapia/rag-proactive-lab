import os
import json
import re
import uuid
import hashlib
from datetime import datetime, timezone
import csv
from typing import Optional
from pathlib import Path

APP_DIR = Path(__file__).resolve().parent
ROOT_DIR = APP_DIR.parent

# ============ Configuración ============
LOG_DIR = os.getenv("LOG_DIR", "logs")
LOG_FILE = os.getenv("LOG_FILE", os.path.join(LOG_DIR, "events.jsonl"))
STORAGE_BACKEND = (os.getenv("STORAGE_BACKEND", "jsonl").strip().lower())  # jsonl | neo4j | dual
MEDIA_QUEUE_FILE = os.getenv("MEDIA_QUEUE_FILE", os.path.join(LOG_DIR, "media_queue.jsonl"))

# Neo4j (opcional)
NEO4J_URI = os.getenv("NEO4J_URI")
NEO4J_USER = os.getenv("NEO4J_USER")
NEO4J_PASSWORD = os.getenv("NEO4J_PASSWORD")
NEO4J_DATABASE = os.getenv("NEO4J_DATABASE", "neo4j")

_neo_driver_cached = None
_neo_constraints_ok = False

# Postgres/Neon (opcional)
POSTGRES_URL = os.getenv("POSTGRES_URL") or os.getenv("NEON_POSTGRES_URL")
_pg_pool_cached = None

# ============ Casos (estado local) ============
CASE_PREFIX = os.getenv("CASE_PREFIX", "AGS").strip() or "AGS"
CASE_TTL_HOURS = int(os.getenv("CASE_TTL_HOURS", "72") or "72")
CASE_STATE_FILE = os.path.join(LOG_DIR, "cases_state.json")
CASE_LOG_FILE = os.path.join(LOG_DIR, "cases.jsonl")
PLAYBOOKS_FILE = os.getenv(
    "PLAYBOOKS_FILE",
    str((ROOT_DIR / "playbooks.json").resolve()),
)


_DTC_CATALOG: dict[str, dict] = {}


def _load_dtc_catalog() -> dict[str, dict]:
    global _DTC_CATALOG
    if _DTC_CATALOG:
        return _DTC_CATALOG
    try:
        catalog_path = (ROOT_DIR / "data" / "dtc_catalog.json").resolve()
        with catalog_path.open("r", encoding="utf-8") as f:
            data = json.load(f)
            if isinstance(data, dict):
                codes = data.get("codes")
                if isinstance(codes, dict):
                    _DTC_CATALOG = {str(k).upper(): v for k, v in codes.items() if isinstance(v, dict)}
    except FileNotFoundError:
        _DTC_CATALOG = {}
    except Exception:
        _DTC_CATALOG = {}
    return _DTC_CATALOG


def _ensure_log_path():
    try:
        os.makedirs(LOG_DIR, exist_ok=True)
    except Exception:
        pass
    # asegurar archivos de casos
    try:
        base = os.path.dirname(CASE_STATE_FILE)
        if base:
            os.makedirs(base, exist_ok=True)
        if not os.path.exists(CASE_STATE_FILE):
            with open(CASE_STATE_FILE, 'w', encoding='utf-8') as f:
                f.write('{}')
    except Exception:
        pass


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def classify(text: str) -> str:
    t = (text or "").lower()
    oem = re.search(r"\b[A-Z0-9]{2,}(?:[\-–][A-Z0-9]{2,})+\b", text or "")
    oil = re.search(r"\boil\s*\d{3}\b", t)
    if any(k in t for k in ["oem", "número de parte", "numero de parte", "refacción", "refaccion", "pieza"]):
        return "parts_lookup"
    if oem:
        return "parts_lookup"
    if any(k in t for k in ["mantenimiento", "servicio", "cada ", "kms", "kilómetros", "kilometros"]):
        return "maintenance"
    if any(k in t for k in ["se apagó", "se apago", "no enciende", "no prende", "se detuvo"]):
        return "engine_issue"
    if any(k in t for k in ["freno", "balata", "abs"]):
        return "brakes"
    if oil:
        return "oil_indicator"
    return "general_question"


def extract_signals(text: str) -> dict:
    t = (text or "").strip()
    tl = t.lower()
    tokens = re.findall(r"\w+", tl)
    dtc_details = []
    dtc_codes: list[str] = []

    # DTC detection (P/B/C/U + four dígitos)
    for code_match in re.findall(r"\b([PBCU][0-9]{4})\b", t.upper()):
        code = code_match.upper()
        if code not in dtc_codes:
            dtc_codes.append(code)

    catalog = _load_dtc_catalog() if dtc_codes else {}
    # Modelo
    model = None
    for m in ["h6c", "h5c", "klq6540"]:
        if m in tokens:
            model = m.upper()
            break
    # OEM
    oem = None
    m_oem = re.search(r"\b[A-Z0-9]{2,}(?:[\-–][A-Z0-9]{2,})+\b", t)
    if m_oem:
        oem = m_oem.group(0)
    # OIL code
    oil = None
    m_oil = re.search(r"\boil\s*(\d{3})\b", t, re.IGNORECASE)
    if m_oil:
        oil = m_oil.group(1)
    # Categoría
    category = None
    if any(k in tl for k in ["freno", "balata", "abs", "liquido de frenos", "líquido de frenos"]):
        category = "brakes"
    elif any(k in tl for k in ["aceite", "oil", "presión de aceite", "presion de aceite"]):
        category = "oil"
    elif any(k in tl for k in ["temperatura", "sobrecalent", "radiador", "anticongelante", "ventilador", "bomba de agua"]):
        category = "cooling"
    elif any(k in tl for k in ["alternador", "fusible", "corto", "cable", "sensor", "ecu", "pcm", "luz de advertencia", "warning", "motor de arranque"]):
        category = "electrical"
    elif any(k in tl for k in ["bateria", "batería"]):
        category = "battery"
    elif any(k in tl for k in ["combustible", "diesel", "gasolina", "cng", "inyector", "bomba", "filtro de combustible", "presión de combustible", "presion de combustible"]):
        category = "fuel"
    elif any(k in tl for k in ["motor", "culata", "bancada"]):
        category = "engine"
    elif any(k in tl for k in ["transmisión", "transmision", "caja", "cambio", "embrague", "diferencial", "eje trasero"]):
        category = "transmission"
    elif any(k in tl for k in ["suspensión", "suspension", "muelle", "brazo", "horquilla"]):
        category = "suspension"
    elif any(k in tl for k in ["dirección", "direccion", "cremallera", "bomba de dirección"]):
        category = "steering"
    elif any(k in tl for k in ["amortiguador", "amortiguadores"]):
        category = "shocks"
    elif any(k in tl for k in ["llanta", "neumático", "neumatico", "neumáticos", "neumaticos", "rin", "rín", "rines"]):
        category = "tires"
    elif any(k in tl for k in ["aire acondicionado", "a/c", "compresor de ac", "condensador", "evaporador", "clima", "calefaccion", "calefacción"]):
        category = "ac"
    elif any(k in tl for k in ["carroceria", "carrocería", "pintura", "corros", "oxido", "óxido"]):
        category = "body"
    elif any(k in tl for k in ["escape", "catalizador", "silenciador"]):
        category = "exhaust"
    # Ajustar categoría con catálogo DTC si aplica
    if not category and dtc_codes:
        for code in dtc_codes:
            info = catalog.get(code) if catalog else None
            if isinstance(info, dict):
                cat = (info.get("category") or "").strip().lower()
                if cat:
                    category = cat
                    break
    # Severidad
    severity = "normal"
    if any(k in tl for k in [
        "sin frenos", "no frena", "freno al fondo", "se apagó", "se apago", "se apaga", "humo", "fuego", "olor a quemado",
        "temperatura alta", "sobrecalent", "presión de aceite", "presion de aceite", "testigo de aceite", "oíl 001", "oil 001", "oil 003",
        "dirección dura", "direccion dura", "no gira", "ruido fuerte", "golpeteo motor", "check engine parpadea"
    ]):
        severity = "critical"
    elif any(k in tl for k in [
        "no enciende", "no arranca", "se jalonea", "pierde potencia", "vibración", "vibracion", "falla intermitente"
    ]):
        severity = "urgent"

    severity_order = {"normal": 0, "urgent": 1, "critical": 2}
    if dtc_codes and catalog:
        for code in dtc_codes:
            info = catalog.get(code)
            if not isinstance(info, dict):
                continue
            dtc_entry = {
                "code": code,
                "description": info.get("description"),
                "category": info.get("category"),
                "severity": info.get("severity"),
            }
            dtc_details.append({k: v for k, v in dtc_entry.items() if v})
            sev = (info.get("severity") or "").strip().lower()
            if sev and severity_order.get(sev, 0) > severity_order.get(severity, 0):
                severity = sev
            cat = (info.get("category") or "").strip().lower()
            if cat and not category:
                category = cat

    # Problema
    problem = None
    if any(k in tl for k in ["fuga", "goteo", "derrama", "chorrea", "burbuja", "burbujas", "espuma", "jabon", "jabón", "burbujeo"]):
        problem = "leak"
    elif any(k in tl for k in ["desgaste", "gastado", "consumido"]):
        problem = "wear"
    elif any(k in tl for k in ["oxid", "corros", "herrumbe"]):
        problem = "oxidation"

    return {
        "model": model,
        "oem": oem,
        "oil_code": oil,
        "length": len(t),
        "category": category,
        "severity": severity,
        "problem": problem,
        "dtc_codes": dtc_codes,
        "dtc_details": dtc_details,
    }


def _neo_driver():
    global _neo_driver_cached
    if _neo_driver_cached is not None:
        return _neo_driver_cached
    if not (NEO4J_URI and NEO4J_USER and NEO4J_PASSWORD):
        return None
    try:
        from neo4j import GraphDatabase
    except Exception:
        return None
    try:
        _neo_driver_cached = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USER, NEO4J_PASSWORD))
        return _neo_driver_cached
    except Exception:
        return None


def _neo_ensure_constraints(session):
    global _neo_constraints_ok
    if _neo_constraints_ok:
        return
    try:
        # Unicidad por id en Event
        session.run("CREATE CONSTRAINT event_id IF NOT EXISTS FOR (e:Event) REQUIRE e.id IS UNIQUE")
        # Unicidad en Contact.address
        session.run("CREATE CONSTRAINT contact_address IF NOT EXISTS FOR (c:Contact) REQUIRE c.address IS UNIQUE")
        # Unicidad de Source por clave sintética
        session.run("CREATE CONSTRAINT source_key IF NOT EXISTS FOR (s:Source) REQUIRE s.key IS UNIQUE")
        _neo_constraints_ok = True
    except Exception:
        pass


def _neo_log_event(kind: str, payload: dict):
    drv = _neo_driver()
    if drv is None:
        return
    event_id = str(uuid.uuid4())
    ts = _now_iso()
    p = payload or {}
    question = p.get("question") or p.get("text")
    answer = p.get("answer")
    endpoint = p.get("endpoint")
    classification = p.get("classification")
    signals = p.get("signals") or {}
    from_addr = p.get("from") or p.get("to")  # guardaremos un nodo Contact si hay dirección
    sources = p.get("sources") or []
    latency_ms = p.get("latency_ms")
    response_bytes = p.get("response_bytes")

    with drv.session(database=NEO4J_DATABASE) as session:
        _neo_ensure_constraints(session)
        # Crear Evento y Contact (si existe)
        session.run(
            """
            MERGE (e:Event {id:$id})
            SET e.ts=$ts, e.kind=$kind, e.endpoint=$endpoint,
                e.question=$question, e.answer=$answer,
                e.classification=$classification, e.signals=$signals,
                e.latency_ms=$latency_ms, e.response_bytes=$response_bytes
            WITH e
            RETURN e
            """,
            id=event_id, ts=ts, kind=kind, endpoint=endpoint,
            question=question, answer=answer,
            classification=classification, signals=signals,
            latency_ms=latency_ms, response_bytes=response_bytes,
        )
        if from_addr:
            session.run(
                """
                MERGE (c:Contact {address:$addr})
                WITH c
                MATCH (e:Event {id:$id})
                MERGE (c)-[:ASSOCIATED_WITH]->(e)
                """,
                addr=str(from_addr), id=event_id,
            )
        # Crear fuentes si hay
        if sources:
            # Prepara lista con key sintética
            srcs = []
            for s in sources:
                doc_id = (s.get('doc_id') if isinstance(s, dict) else None) or 'SSOT-HIGER'
                page_label = str((s.get('page_label') if isinstance(s, dict) else 'N/A'))
                chunk_index = s.get('chunk_index') if isinstance(s, dict) else None
                key = f"{doc_id}|{page_label}|{chunk_index if chunk_index is not None else -1}"
                srcs.append({
                    'key': key,
                    'doc_id': doc_id,
                    'page_label': page_label,
                    'page_number': s.get('page_number') if isinstance(s, dict) else None,
                    'chunk_index': chunk_index if chunk_index is not None else -1,
                    'snippet': s.get('snippet') if isinstance(s, dict) else None,
                })
            session.run(
                """
                MATCH (e:Event {id:$id})
                UNWIND $sources AS s
                MERGE (src:Source {key:s.key})
                SET src.doc_id=s.doc_id, src.page_label=s.page_label,
                    src.page_number=s.page_number, src.chunk_index=s.chunk_index,
                    src.snippet=s.snippet
                MERGE (e)-[:HAS_SOURCE]->(src)
                """,
                id=event_id, sources=srcs,
            )


def _jsonl_log_event(kind: str, payload: dict):
    _ensure_log_path()
    event = {"ts": _now_iso(), "kind": kind, **(payload or {})}
    try:
        with open(LOG_FILE, 'a', encoding='utf-8') as f:
            f.write(json.dumps(event, ensure_ascii=False) + "\n")
    except Exception:
        pass


def _pg_pool():
    global _pg_pool_cached
    if _pg_pool_cached is not None:
        return _pg_pool_cached
    if not POSTGRES_URL:
        return None
    try:
        from psycopg_pool import ConnectionPool  # type: ignore
    except Exception:
        return None
    try:
        _pg_pool_cached = ConnectionPool(conninfo=POSTGRES_URL, max_size=5)
        return _pg_pool_cached
    except Exception:
        return None


def _pg_ensure_schema(conn):
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS events (
                  id UUID PRIMARY KEY,
                  ts TIMESTAMPTZ NOT NULL,
                  kind TEXT NOT NULL,
                  endpoint TEXT,
                  question TEXT,
                  answer TEXT,
                  classification TEXT,
                  signals JSONB,
                  meta JSONB,
                  contact_address TEXT,
                  conversation_id TEXT,
                  channel TEXT,
                  latency_ms DOUBLE PRECISION,
                  response_bytes BIGINT
                );
                CREATE TABLE IF NOT EXISTS sources (
                  event_id UUID REFERENCES events(id) ON DELETE CASCADE,
                  doc_id TEXT,
                  page_label TEXT,
                  page_number INT,
                  chunk_index INT,
                  snippet TEXT,
                  PRIMARY KEY (event_id, doc_id, page_label, chunk_index)
                );
                CREATE INDEX IF NOT EXISTS idx_events_ts ON events(ts DESC);
                CREATE INDEX IF NOT EXISTS idx_events_kind ON events(kind);
                CREATE INDEX IF NOT EXISTS idx_events_contact ON events(contact_address);
                ALTER TABLE events ADD COLUMN IF NOT EXISTS conversation_id TEXT;
                ALTER TABLE events ADD COLUMN IF NOT EXISTS channel TEXT;
                ALTER TABLE events ADD COLUMN IF NOT EXISTS latency_ms DOUBLE PRECISION;
                ALTER TABLE events ADD COLUMN IF NOT EXISTS response_bytes BIGINT;
                """
            )
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass


def _pg_log_event(kind: str, payload: dict):
    if not POSTGRES_URL:
        return
    try:
        import psycopg  # type: ignore
        from psycopg.types.json import Json  # type: ignore
    except Exception:
        return

    event_id = str(uuid.uuid4())
    ts = _now_iso()
    p = payload or {}
    endpoint = p.get("endpoint")
    question = p.get("question") or p.get("text")
    answer = p.get("answer")
    classification = p.get("classification")
    signals = p.get("signals") or {}
    meta = p.get("meta") or {}
    contact_addr = p.get("from") or p.get("to") or (meta.get("contact") if isinstance(meta, dict) else None)
    conversation_id = p.get("conversation_id") or (meta.get("session_id") if isinstance(meta, dict) else None)
    channel = p.get("channel") or (meta.get("channel") if isinstance(meta, dict) else None)
    sources = p.get("sources") or []
    latency_ms = p.get("latency_ms")
    response_bytes = p.get("response_bytes")

    pool = _pg_pool()
    if pool is not None:
        try:
            with pool.connection() as conn:
                _pg_ensure_schema(conn)
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        INSERT INTO events (
                            id, ts, kind, endpoint, question, answer, classification, signals, meta,
                            contact_address, conversation_id, channel, latency_ms, response_bytes
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        (
                            event_id,
                            ts,
                            kind,
                            endpoint,
                            question,
                            answer,
                            classification,
                            Json(signals),
                            Json(meta),
                            str(contact_addr) if contact_addr else None,
                            conversation_id,
                            channel,
                            float(latency_ms) if latency_ms is not None else None,
                            int(response_bytes) if response_bytes is not None else None,
                        ),
                    )
                    if sources:
                        for s in sources:
                            cur.execute(
                                """
                                INSERT INTO sources (event_id, doc_id, page_label, page_number, chunk_index, snippet)
                                VALUES (%s, %s, %s, %s, %s, %s)
                                ON CONFLICT (event_id, doc_id, page_label, chunk_index) DO NOTHING
                                """,
                                (
                                    event_id,
                                    (s.get('doc_id') if isinstance(s, dict) else None) or 'SSOT-HIGER',
                                    str((s.get('page_label') if isinstance(s, dict) else 'N/A')),
                                    s.get('page_number') if isinstance(s, dict) else None,
                                    (s.get('chunk_index') if isinstance(s, dict) else None) or -1,
                                    s.get('snippet') if isinstance(s, dict) else None,
                                ),
                            )
                conn.commit()
            return
        except Exception:
            pass

    # Sin pool: conexión directa por evento
    try:
        with psycopg.connect(POSTGRES_URL, autocommit=True) as conn:
            _pg_ensure_schema(conn)
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO events (
                        id, ts, kind, endpoint, question, answer, classification, signals, meta,
                        contact_address, conversation_id, channel, latency_ms, response_bytes
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        event_id,
                        ts,
                        kind,
                        endpoint,
                        question,
                        answer,
                        classification,
                        Json(signals),
                        Json(meta),
                        str(contact_addr) if contact_addr else None,
                        conversation_id,
                        channel,
                        float(latency_ms) if latency_ms is not None else None,
                        int(response_bytes) if response_bytes is not None else None,
                    ),
                )
                if sources:
                    for s in sources:
                        cur.execute(
                            """
                            INSERT INTO sources (event_id, doc_id, page_label, page_number, chunk_index, snippet)
                            VALUES (%s, %s, %s, %s, %s, %s)
                            ON CONFLICT (event_id, doc_id, page_label, chunk_index) DO NOTHING
                            """,
                            (
                                event_id,
                                (s.get('doc_id') if isinstance(s, dict) else None) or 'SSOT-HIGER',
                                str((s.get('page_label') if isinstance(s, dict) else 'N/A')),
                                s.get('page_number') if isinstance(s, dict) else None,
                                (s.get('chunk_index') if isinstance(s, dict) else None) or -1,
                                s.get('snippet') if isinstance(s, dict) else None,
                            ),
                        )
    except Exception:
        pass


def log_event(kind: str, payload: dict):
    backend = STORAGE_BACKEND
    # Siempre guardamos JSONL como respaldo a menos que se pida solo neo4j
    if backend in ("jsonl", "dual", ""):
        _jsonl_log_event(kind, payload)
    if backend in ("neo4j", "dual") or (backend == "jsonl" and (NEO4J_URI and NEO4J_USER and NEO4J_PASSWORD)):
        _neo_log_event(kind, payload)
    if backend in ("postgres", "pg", "dual") or (backend == "jsonl" and POSTGRES_URL):
        _pg_log_event(kind, payload)


def get_conversation(contact: str, limit: int = 8):
    """Devuelve historial breve de conversación como lista de dicts [{role: 'user'|'assistant', text: str}].
    Usa Postgres si está configurado; si no, intenta JSONL; si no, vacío.
    """
    if not contact:
        return []
    # Postgres primero
    if POSTGRES_URL:
        try:
            import psycopg  # type: ignore
            q = (
                "SELECT ts, kind, question, answer FROM events "
                "WHERE contact_address = %s ORDER BY ts DESC LIMIT %s"
            )
            with psycopg.connect(POSTGRES_URL) as conn:
                with conn.cursor() as cur:
                    cur.execute(q, (str(contact), int(limit)))
                    rows = cur.fetchall()
            # Construir en orden cronológico
            rows = list(reversed(rows))
            msgs = []
            for (_ts, kind, qtext, atext) in rows:
                k = (kind or '').lower()
                if k in ("whatsapp_in", "api_query") and qtext:
                    msgs.append({"role": "user", "text": qtext})
                if k in ("whatsapp_out", "api_query") and atext:
                    msgs.append({"role": "assistant", "text": atext})
            return msgs[-limit:]
        except Exception:
            pass
    # JSONL fallback
    try:
        path = LOG_FILE
        if not (path and os.path.exists(path)):
            return []
        items = []
        with open(path, 'r', encoding='utf-8') as f:
            for line in f:
                try:
                    ev = json.loads(line)
                except Exception:
                    continue
                addr = ev.get('from') or ev.get('to') or ((ev.get('meta') or {}).get('contact') if isinstance(ev.get('meta'), dict) else None)
                if not addr or str(addr) != str(contact):
                    continue
                k = (ev.get('kind') or '').lower()
                qtext = ev.get('question') or ev.get('text')
                atext = ev.get('answer')
                if k in ("whatsapp_in", "api_query") and qtext:
                    items.append({"role": "user", "text": qtext})
                if k in ("whatsapp_out", "api_query") and atext:
                    items.append({"role": "assistant", "text": atext})
        return items[-limit:]
    except Exception:
        return []


# ---------------------- Playbooks + Casos ----------------------
_playbooks_cache = None


def load_playbooks() -> dict:
    global _playbooks_cache
    if _playbooks_cache is not None:
        return _playbooks_cache
    pb = None
    # 1) archivo externo
    try:
        if PLAYBOOKS_FILE and os.path.exists(PLAYBOOKS_FILE):
            with open(PLAYBOOKS_FILE, 'r', encoding='utf-8') as f:
                pb = json.load(f)
    except Exception:
        pb = None
    # 2) default mínimo
    if not isinstance(pb, dict):
        pb = {
            "brakes": {
                "immediate_steps": [
                    "Detén la unidad si el pedal se va al fondo",
                    "Verifica nivel de líquido de frenos",
                ],
                "short_checklist": [
                    "Revisa fugas en latiguillos y calipers",
                    "Purgar el sistema si hay aire",
                    "Revisar servofreno y líneas",
                ],
                "ask_for": ["foto_vin", "foto_placa", "foto_odometro", "fotos_fugas_frenos"],
                "route_suggestion": "taller"
            },
            "oil": {
                "immediate_steps": [
                    "Detén la unidad si el testigo de aceite está encendido",
                    "Verifica nivel de aceite en frío y superficie nivelada",
                ],
                "short_checklist": [
                    "Revisar fugas en cárter/filtro",
                    "Confirmar tipo/viscosidad de aceite",
                ],
                "ask_for": ["foto_testigo_aceite", "foto_varilla_aceite", "foto_odometro", "foto_vin"],
                "route_suggestion": "taller"
            },
            "cooling": {
                "immediate_steps": [
                    "Detén la unidad si hay sobrecalentamiento",
                    "Espera a que enfríe antes de abrir depósito",
                ],
                "short_checklist": [
                    "Nivel en depósito",
                    "Fugas visibles en mangueras/radiador",
                    "Funcionamiento del ventilador",
                ],
                "ask_for": ["foto_vin", "foto_odometro", "foto_tablero_temp", "fotos_fugas_cooling"],
                "route_suggestion": "taller"
            },
            "electrical": {
                "immediate_steps": [
                    "Apaga el sistema si hay olor a quemado/humo",
                ],
                "short_checklist": [
                    "Fotos del tablero y fusibles",
                    "Revisar bornes y tierra",
                ],
                "ask_for": ["foto_tablero", "foto_fusibles", "foto_vin", "foto_odometro"],
                "route_suggestion": "taller"
            },
            "fuel": {
                "immediate_steps": [
                    "No forzar marcha si hay jaloneos fuertes",
                ],
                "short_checklist": [
                    "Filtro de combustible (fecha)",
                    "Presión de combustible",
                ],
                "ask_for": ["foto_odometro", "video_falla", "foto_vin"],
                "route_suggestion": "taller"
            },
            "general": {
                "immediate_steps": [],
                "short_checklist": [],
                "ask_for": ["foto_vin", "foto_odometro"],
                "route_suggestion": "evaluar"
            }
        }
    _playbooks_cache = pb
    return pb


def _load_cases_state() -> dict:
    try:
        if os.path.exists(CASE_STATE_FILE):
            with open(CASE_STATE_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if isinstance(data, dict):
                    return data
    except Exception:
        pass
    return {}


def _save_cases_state(state: dict):
    try:
        with open(CASE_STATE_FILE, 'w', encoding='utf-8') as f:
            json.dump(state, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _append_case_log(entry: dict):
    try:
        with open(CASE_LOG_FILE, 'a', encoding='utf-8') as f:
            f.write(json.dumps(entry, ensure_ascii=False) + "\n")
    except Exception:
        pass


def _new_case_id() -> str:
    d = datetime.now(timezone.utc).strftime('%Y%m%d')
    suffix = uuid.uuid4().hex[:4].upper()
    return f"{CASE_PREFIX}-{d}-{suffix}"


def get_or_create_case(contact: str) -> dict:
    _ensure_log_path()
    state = _load_cases_state()
    c = state.get(contact)
    now = datetime.now(timezone.utc).isoformat()
    if not c:
        c = {
            'id': _new_case_id(),
            'contact': contact,
            'opened_at': now,
            'last_updated': now,
            'severity': 'normal',
            'categories': [],
            'required': [],
            'provided': [],
            'required_log': {},
            'provided_log': {},
            'attachments': [],
            'status': 'open',
        }
        state[contact] = c
        _save_cases_state(state)
        _append_case_log({'ts': now, 'op': 'create', **c})
    else:
        # Compatibilidad: asegurar campos recientes
        c.setdefault('required_log', {})
        c.setdefault('provided_log', {})
        c.setdefault('attachments', [])
    return c


def update_case(contact: str, patch: dict) -> dict:
    state = _load_cases_state()
    c = state.get(contact) or get_or_create_case(contact)
    _ensure_case_logs(c)
    for k, v in (patch or {}).items():
        if k == 'categories' and isinstance(v, list):
            cur = c.get('categories') or []
            merged = list(dict.fromkeys(cur + v))
            c['categories'] = merged
        elif k == 'required' and isinstance(v, list):
            _merge_required(c, v)
        elif k == 'provided' and isinstance(v, list):
            _merge_provided(c, v)
        elif k == 'attachments' and isinstance(v, list):
            c['attachments'] = _merge_attachments(c.get('attachments') or [], v)
        elif v is not None:
            c[k] = v
    _sync_requirement_fulfillment(c)
    c['last_updated'] = datetime.now(timezone.utc).isoformat()
    state[contact] = c
    _save_cases_state(state)
    _append_case_log({'ts': c['last_updated'], 'op': 'update', 'contact': contact, 'patch': patch})
    return c


def _ensure_case_logs(case: dict) -> None:
    case.setdefault('required_log', {})
    case.setdefault('provided_log', {})
    case.setdefault('attachments', [])


def _merge_required(case: dict, items: list[str]) -> None:
    now = _now_iso()
    req = case.get('required') or []
    log = case.get('required_log') or {}
    for item in items:
        norm = (item or '').strip()
        if not norm:
            continue
        if norm not in req:
            req.append(norm)
            entry = log.get(norm, {})
            entry['first_requested_at'] = entry.get('first_requested_at') or now
        else:
            entry = log.get(norm, {})
            entry['first_requested_at'] = entry.get('first_requested_at') or now
        entry['last_requested_at'] = now
        log[norm] = entry
    case['required'] = list(dict.fromkeys(req))
    case['required_log'] = log


def _merge_provided(case: dict, items: list[str]) -> None:
    now = _now_iso()
    provided = case.get('provided') or []
    log = case.get('provided_log') or {}
    for item in items:
        norm = (item or '').strip()
        if not norm:
            continue
        if norm not in provided:
            provided.append(norm)
        entry = log.get(norm, {})
        entry['first_provided_at'] = entry.get('first_provided_at') or now
        entry['last_provided_at'] = now
        log[norm] = entry
    case['provided'] = list(dict.fromkeys(provided))
    case['provided_log'] = log


def _attachment_hash_local(url: str, object_key: Optional[str] = None) -> Optional[str]:
    base = (url or '').strip()
    if not base:
        return None
    extra = (object_key or '').strip()
    payload = base if not extra else f"{base}|{extra}"
    return hashlib.sha256(payload.encode('utf-8', 'ignore')).hexdigest()


def _merge_attachments(existing: list[dict], incoming: list[dict]) -> list[dict]:
    now = _now_iso()
    merged: list[dict] = []
    index: dict[str, dict] = {}

    def _add_existing(item: dict) -> None:
        url = (item.get('url') or '').strip()
        if not url:
            return
        key = item.get('url_hash') or _attachment_hash_local(url, item.get('object_key'))
        if not key:
            return
        item['url_hash'] = key
        item.setdefault('first_seen_at', item.get('first_seen_at') or now)
        item.setdefault('last_seen_at', item.get('last_seen_at') or item.get('first_seen_at') or now)
        merged.append(item)
        index[key] = item

    for item in existing or []:
        if isinstance(item, dict):
            _add_existing(dict(item))

    for item in incoming or []:
        if not isinstance(item, dict):
            continue
        url = (item.get('url') or '').strip()
        if not url:
            continue
        key = _attachment_hash_local(url, item.get('object_key'))
        if not key:
            continue
        if key in index:
            entry = index[key]
            entry['last_seen_at'] = now
            if not entry.get('content_type') and item.get('content_type'):
                entry['content_type'] = item.get('content_type')
        else:
            entry = {
                'url': url,
                'content_type': item.get('content_type'),
                'object_key': item.get('object_key'),
                'url_hash': key,
                'first_seen_at': now,
                'last_seen_at': now,
            }
            merged.append(entry)
            index[key] = entry
    return merged


def _sync_requirement_fulfillment(case: dict) -> None:
    now = _now_iso()
    required_log = case.get('required_log') or {}
    provided_log = case.get('provided_log') or {}
    provided_set = set(case.get('provided') or [])
    for item, entry in required_log.items():
        if item in provided_set:
            fulfilled = entry.get('fulfilled_at')
            if not fulfilled:
                fulfilled_at = provided_log.get(item, {}).get('first_provided_at') or now
                entry['fulfilled_at'] = fulfilled_at
    case['required_log'] = required_log


def add_required(contact: str, items: list[str]):
    update_case(contact, {'required': items})


def mark_provided(contact: str, items: list[str]):
    update_case(contact, {'provided': items})


def attach_media(contact: str, media_list: list[dict]):
    if not media_list:
        return
    safe = []
    for m in media_list:
        url = m.get('url') if isinstance(m, dict) else None
        ctype = m.get('content_type') if isinstance(m, dict) else None
        if url:
            safe.append({'url': url, 'content_type': ctype})
    if safe:
        update_case(contact, {'attachments': safe})


def enqueue_media(contact: str, media_item: dict):
    if not contact or not isinstance(media_item, dict):
        return
    _ensure_log_path()
    payload = {
        'ts': _now_iso(),
        'contact': contact,
        'media': {
            'url': media_item.get('url'),
            'content_type': media_item.get('content_type'),
            'object_key': media_item.get('object_key'),
        },
    }
    try:
        with open(MEDIA_QUEUE_FILE, 'a', encoding='utf-8') as f:
            f.write(json.dumps(payload, ensure_ascii=False) + "\n")
    except Exception:
        pass


def get_case_state(contact: str) -> dict | None:
    if not contact:
        return None
    state = _load_cases_state()
    case = state.get(contact)
    if not case:
        return None
    return dict(case)


def case_requirements_snapshot(contact: str, case: dict | None = None) -> dict:
    snapshot = {'required': [], 'provided': [], 'missing': []}
    case = case or get_case_state(contact)
    if not case:
        return snapshot
    _ensure_case_logs(case)
    required = list(dict.fromkeys((case.get('required') or [])))
    provided = list(dict.fromkeys((case.get('provided') or [])))
    try:
        pb = load_playbooks()
        for cat in case.get('categories') or []:
            ask_for = (pb.get(cat) or {}).get('ask_for') or []
            for item in ask_for:
                if item not in required:
                    required.append(item)
    except Exception:
        pass
    provided_set = set(provided)
    missing = [item for item in required if item not in provided_set]
    snapshot['required'] = required
    snapshot['provided'] = provided
    snapshot['missing'] = missing
    required_log = case.get('required_log') or {}
    provided_log = case.get('provided_log') or {}
    snapshot['required_details'] = [
        {
            'item': item,
            'first_requested_at': (required_log.get(item) or {}).get('first_requested_at'),
            'last_requested_at': (required_log.get(item) or {}).get('last_requested_at'),
            'fulfilled_at': (required_log.get(item) or {}).get('fulfilled_at'),
        }
        for item in required
    ]
    snapshot['provided_details'] = [
        {
            'item': item,
            'first_provided_at': (provided_log.get(item) or {}).get('first_provided_at'),
            'last_provided_at': (provided_log.get(item) or {}).get('last_provided_at'),
        }
        for item in provided
    ]
    return snapshot


def list_cases() -> list[dict]:
    state = _load_cases_state()
    summaries = []
    for contact, case in state.items():
        snapshot = case_requirements_snapshot(contact, case)
        summaries.append({
            'contact': contact,
            'id': case.get('id'),
            'severity': case.get('severity'),
            'categories': case.get('categories'),
            'status': case.get('status'),
            'opened_at': case.get('opened_at'),
            'last_updated': case.get('last_updated'),
            'missing': snapshot.get('missing', []),
            'missing_count': len(snapshot.get('missing') or []),
        })
    summaries.sort(key=lambda item: item.get('last_updated') or '', reverse=True)
    return summaries


def export_csv(events_csv_path: str | None = None, sources_csv_path: str | None = None, prefer_postgres: bool = True):
    """
    Exporta eventos y fuentes a CSV.
    - Si hay Postgres (Neon) y prefer_postgres=True, exporta desde BD.
    - Si no, cae a JSONL local.
    """
    events_csv_path = events_csv_path or os.path.join(LOG_DIR, 'events.csv')
    sources_csv_path = sources_csv_path or os.path.join(LOG_DIR, 'sources.csv')
    os.makedirs(os.path.dirname(events_csv_path), exist_ok=True)

    def _dump_events(rows, out_path):
        cols = [
            'id','ts','kind','endpoint','question','answer','classification',
            'signals','meta','contact_address','conversation_id','channel'
        ]
        with open(out_path, 'w', newline='', encoding='utf-8') as f:
            w = csv.DictWriter(f, fieldnames=cols)
            w.writeheader()
            for r in rows:
                w.writerow({k: (json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v) for k, v in r.items() if k in cols})

    def _dump_sources(rows, out_path):
        cols = ['event_id','doc_id','page_label','page_number','chunk_index','snippet']
        with open(out_path, 'w', newline='', encoding='utf-8') as f:
            w = csv.DictWriter(f, fieldnames=cols)
            w.writeheader()
            for r in rows:
                w.writerow(r)

    # 1) Intentar Postgres
    if prefer_postgres and POSTGRES_URL:
        try:
            import psycopg  # type: ignore
            with psycopg.connect(POSTGRES_URL) as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT id, ts, kind, endpoint, question, answer, classification, signals, meta, contact_address, conversation_id, channel FROM events ORDER BY ts ASC"
                    )
                    ev_rows_raw = cur.fetchall()
                    events_rows = []
                    for (id_, ts, kind, endpoint, q, a, cls, sig, meta, contact, conv, ch) in ev_rows_raw:
                        events_rows.append({
                            'id': str(id_), 'ts': ts.isoformat() if hasattr(ts, 'isoformat') else ts,
                            'kind': kind, 'endpoint': endpoint, 'question': q, 'answer': a,
                            'classification': cls, 'signals': sig, 'meta': meta,
                            'contact_address': contact, 'conversation_id': conv, 'channel': ch,
                        })
                    cur.execute(
                        "SELECT event_id, doc_id, page_label, page_number, chunk_index, snippet FROM sources"
                    )
                    src_rows_raw = cur.fetchall()
                    sources_rows = []
                    for (eid, doc, pl, pn, ci, snip) in src_rows_raw:
                        sources_rows.append({
                            'event_id': str(eid), 'doc_id': doc, 'page_label': pl,
                            'page_number': pn, 'chunk_index': ci, 'snippet': snip,
                        })
            _dump_events(events_rows, events_csv_path)
            _dump_sources(sources_rows, sources_csv_path)
            return {'events_csv': events_csv_path, 'sources_csv': sources_csv_path, 'source': 'postgres'}
        except Exception:
            pass

    # 2) Fallback JSONL
    items = []
    try:
        if LOG_FILE and os.path.exists(LOG_FILE):
            with open(LOG_FILE, 'r', encoding='utf-8') as f:
                for line in f:
                    try:
                        items.append(json.loads(line))
                    except Exception:
                        continue
    except Exception:
        items = []
    events_rows = []
    sources_rows = []
    for ev in items:
        # Generar ID estable desde ts/kind/question
        base = f"{ev.get('ts')}|{ev.get('kind')}|{(ev.get('question') or ev.get('text') or '')}"
        eid = str(uuid.uuid5(uuid.NAMESPACE_URL, base))
        events_rows.append({
            'id': eid,
            'ts': ev.get('ts'),
            'kind': ev.get('kind'),
            'endpoint': ev.get('endpoint'),
            'question': ev.get('question') or ev.get('text'),
            'answer': ev.get('answer'),
            'classification': ev.get('classification'),
            'signals': ev.get('signals'),
            'meta': ev.get('meta'),
            'contact_address': ev.get('from') or ev.get('to') or ((ev.get('meta') or {}).get('contact') if isinstance(ev.get('meta'), dict) else None),
            'conversation_id': (ev.get('meta') or {}).get('session_id') if isinstance(ev.get('meta'), dict) else None,
            'channel': (ev.get('meta') or {}).get('channel') if isinstance(ev.get('meta'), dict) else None,
        })
        for s in (ev.get('sources') or []):
            sources_rows.append({
                'event_id': eid,
                'doc_id': (s.get('doc_id') if isinstance(s, dict) else None) or 'SSOT-HIGER',
                'page_label': str((s.get('page_label') if isinstance(s, dict) else 'N/A')),
                'page_number': s.get('page_number') if isinstance(s, dict) else None,
                'chunk_index': (s.get('chunk_index') if isinstance(s, dict) else None) or -1,
                'snippet': s.get('snippet') if isinstance(s, dict) else None,
            })
    _dump_events(events_rows, events_csv_path)
    _dump_sources(sources_rows, sources_csv_path)
    return {'events_csv': events_csv_path, 'sources_csv': sources_csv_path, 'source': 'jsonl'}


def export_xlsx(xlsx_path: Optional[str] = None, prefer_postgres: bool = True):
    """
    Exporta eventos y fuentes a un Excel (.xlsx) con dos hojas: events y sources.
    Usa Postgres si está disponible (y prefer_postgres=True), de lo contrario JSONL local.
    """
    # Reutilizar la recolección de filas de export_csv, sin escribir CSVs.
    events_rows = []
    sources_rows = []

    # 1) Intentar Postgres
    used = 'jsonl'
    if prefer_postgres and POSTGRES_URL:
        try:
            import psycopg  # type: ignore
            with psycopg.connect(POSTGRES_URL) as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT id, ts, kind, endpoint, question, answer, classification, signals, meta, contact_address, conversation_id, channel FROM events ORDER BY ts ASC"
                    )
                    ev_rows_raw = cur.fetchall()
                    for (id_, ts, kind, endpoint, q, a, cls, sig, meta, contact, conv, ch) in ev_rows_raw:
                        events_rows.append({
                            'id': str(id_), 'ts': ts.isoformat() if hasattr(ts, 'isoformat') else ts,
                            'kind': kind, 'endpoint': endpoint, 'question': q, 'answer': a,
                            'classification': cls, 'signals': sig, 'meta': meta,
                            'contact_address': contact, 'conversation_id': conv, 'channel': ch,
                        })
                    cur.execute(
                        "SELECT event_id, doc_id, page_label, page_number, chunk_index, snippet FROM sources"
                    )
                    src_rows_raw = cur.fetchall()
                    for (eid, doc, pl, pn, ci, snip) in src_rows_raw:
                        sources_rows.append({
                            'event_id': str(eid), 'doc_id': doc, 'page_label': pl,
                            'page_number': pn, 'chunk_index': ci, 'snippet': snip,
                        })
            used = 'postgres'
        except Exception:
            events_rows = []
            sources_rows = []

    # 2) Fallback JSONL
    if not events_rows:
        try:
            items = []
            if LOG_FILE and os.path.exists(LOG_FILE):
                with open(LOG_FILE, 'r', encoding='utf-8') as f:
                    for line in f:
                        try:
                            items.append(json.loads(line))
                        except Exception:
                            continue
            for ev in items:
                base = f"{ev.get('ts')}|{ev.get('kind')}|{(ev.get('question') or ev.get('text') or '')}"
                eid = str(uuid.uuid5(uuid.NAMESPACE_URL, base))
                events_rows.append({
                    'id': eid,
                    'ts': ev.get('ts'),
                    'kind': ev.get('kind'),
                    'endpoint': ev.get('endpoint'),
                    'question': ev.get('question') or ev.get('text'),
                    'answer': ev.get('answer'),
                    'classification': ev.get('classification'),
                    'signals': ev.get('signals'),
                    'meta': ev.get('meta'),
                    'contact_address': ev.get('from') or ev.get('to') or ((ev.get('meta') or {}).get('contact') if isinstance(ev.get('meta'), dict) else None),
                    'conversation_id': (ev.get('meta') or {}).get('session_id') if isinstance(ev.get('meta'), dict) else None,
                    'channel': (ev.get('meta') or {}).get('channel') if isinstance(ev.get('meta'), dict) else None,
                })
                for s in (ev.get('sources') or []):
                    sources_rows.append({
                        'event_id': eid,
                        'doc_id': (s.get('doc_id') if isinstance(s, dict) else None) or 'SSOT-HIGER',
                        'page_label': str((s.get('page_label') if isinstance(s, dict) else 'N/A')),
                        'page_number': s.get('page_number') if isinstance(s, dict) else None,
                        'chunk_index': (s.get('chunk_index') if isinstance(s, dict) else None) or -1,
                        'snippet': s.get('snippet') if isinstance(s, dict) else None,
                    })
            used = 'jsonl'
        except Exception:
            pass

    # 3) Escribir XLSX
    xlsx_path = xlsx_path or os.path.join(LOG_DIR, 'events_and_sources.xlsx')
    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)
    from openpyxl import Workbook  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore

    wb = Workbook()
    # Sheet 1: events
    ws1 = wb.active
    ws1.title = 'events'
    ev_cols = ['id','ts','kind','endpoint','question','answer','classification','signals','meta','contact_address','conversation_id','channel']
    ws1.append(ev_cols)
    for r in events_rows:
        row = []
        for c in ev_cols:
            v = r.get(c)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False)
            row.append(v)
        ws1.append(row)
    # Auto ancho básico
    for idx, col in enumerate(ev_cols, start=1):
        ws1.column_dimensions[get_column_letter(idx)].width = min(max(len(col), 16), 40)

    # Sheet 2: sources
    ws2 = wb.create_sheet('sources')
    src_cols = ['event_id','doc_id','page_label','page_number','chunk_index','snippet']
    ws2.append(src_cols)
    for r in sources_rows:
        ws2.append([r.get(c) for c in src_cols])
    for idx, col in enumerate(src_cols, start=1):
        ws2.column_dimensions[get_column_letter(idx)].width = min(max(len(col), 12), 40)

    wb.save(xlsx_path)
    return {'xlsx': xlsx_path, 'source': used}
